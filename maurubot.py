from selenium import webdriver
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.common.keys import Keys
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from tkinter import Tk
from tkinter.filedialog import askopenfilename
import re, os, sys
from datetime import datetime
from time import time, sleep
from random import randint
from urllib.parse import urlparse
# from   urllib.request                     import urlopen
import xml.etree.ElementTree              as     ET

IGNORE_OPTIONS = {0: 'start', 1: 'within', 2: 'end'}


BOTNAME = 'Maurubot'

GOOG_SEARCH_LINK = 'https://www.google.com/search?q='
BING_SEARCH_LINK = 'https://www.bing.com/search?q='

GOOG_START_STRING = '&start='
BING_START_STRING = '&first='

HEADLESS_OPTION = 'headless'

CHROME_VERSION_PAGE = 'chrome://version/'

CONFIG_FILE = './XML/config.xml'
IGNORE_FILE = './XML/ignore.xml'


def set_up_variables():
    """
    Sets up main variables from XML configuration files
    """
    # load ignore list
    IGNORE_LIST = []
    try:
        items = ET.parse(IGNORE_FILE).getroot().findall('item')
        for item in items:
            IGNORE_LIST.append((item.text.strip(), int(item.attrib['option'])))
        # print(IGNORE_LIST)
        print("Ignore list loaded")
    except:
        print('WARNING: could not load the ignore list. Using default ignore list instead')
        print(sys.exc_info())
        IGNORE_LIST = [('quora.', 1), ('.google.', 1), ('/google.', 1), ('#', 0), ('webcache.googleusercontent', 1),
                       ('facebook.', 1), ('youtube.', 1), ('tumblr.', 1), ('wikipedia.', 1), ('/search', 0)]
    try:
        DEPTH = int(ET.parse(CONFIG_FILE).getroot().findall('depth')[0].text.strip())
        if DEPTH < 0:
            print('WARNING: depth should be larger than 0. Using default depth value 10')
            DEPTH = 10
        print('DEPTH ' + str(DEPTH))
    except:
        print('WARNING: could not load the depth. Using default depth value 10')
        print(sys.exc_info())
        DEPTH = 10
    try:
        EXPORT_THRESHOLD = int(ET.parse(CONFIG_FILE).getroot().findall('export_threshold')[0].text.strip())
        if EXPORT_THRESHOLD < 0:
            print('WARNING: the export threshold should be larger than 0. Using default threshold value 10')
            EXPORT_THRESHOLD = 10
        print('EXPORT_THRESHOLD ' + str(EXPORT_THRESHOLD))
    except:
        print('WARNING: could not load the export threshold. Using default threshold value 10')
        print(sys.exc_info())
        EXPORT_THRESHOLD = 10

    return IGNORE_LIST, DEPTH, EXPORT_THRESHOLD


IGNORE_LIST, DEPTH, EXPORT_THRESHOLD = set_up_variables()


def set_up_driver():
    """
    Sets up driver options and object
    """
    options = webdriver.ChromeOptions()
    CHROME_BINARY = ET.parse(CONFIG_FILE).getroot().findall('chrome')[0].text.strip()
    options.binary_location = CHROME_BINARY  # 'C:\\Program Files (x86)\\Google\\Chrome\\Application\\chrome.exe'
    driver = webdriver.Chrome(options=options)
    chrome_page = driver.get(CHROME_VERSION_PAGE)
    soup = BeautifulSoup(driver.page_source, 'lxml')

    default_profile = None

    tds = soup.find_all('td')

    for td in tds:
        try:
            if td.get('id') == 'profile_path':
                default_profile = td.get_text()

                options.add_argument("user-data-dir=" + default_profile)
                print("Default profile path loaded")
        except:
            print("WARNING: Could not load default profile path")
            print(sys.exc_info())
            continue

    options.add_argument(HEADLESS_OPTION)
    driver.maximize_window()
    driver = webdriver.Chrome(options=options)

    return driver


def adjust_search_string(search_string):
    """
    """
    print("Adjusting search string " + search_string)
    ss_split = search_string.split()
    ss_attach = ss_split[0]
    for i in range(1, len(ss_split)):
        ss_attach += '+' + ss_split[i]
    return ss_attach


def load_keyword_dict(filepath):
    """
    """
    wb = load_workbook(filepath)
    sheet = wb.active
    keywords = {}
    for i in range(sheet.max_row):
        # print(sheet['A'+str(i+1)].value)
        cell_value = sheet['A' + str(i + 1)].value
        if cell_value is not None:
            status = sheet['B' + str(i + 1)].value
            keywords[cell_value] = status

    return keywords


def set_keyword_status(filepath, keyword, status):
    """
    """
    try:
        wb = load_workbook(filepath)
        sheet = wb.active
        for i in range(sheet.max_row):
            # print(sheet['A'+str(i+1)].value)
            cell_value = sheet['A' + str(i + 1)].value
            if cell_value is not None and cell_value == keyword:
                sheet['B' + str(i + 1)] = status
                wb.save(filepath)
                return True
    except:
        print('Error when trying to change keyword status')
        print(sys.exc_info())
        return False
    return False


def ignore_single_link(link):
    for ignore_par in IGNORE_LIST:
        if ignore_par[1] == 0:
            if link[:len(ignore_par[0])] == ignore_par[0]:
                print(link + ' ignored, with option ' + IGNORE_OPTIONS[ignore_par[1]] + ' for ' + ignore_par[0])
                return True
        elif ignore_par[1] == 1:
            if ignore_par[0] in link:
                print(link + ' ignored, with option ' + IGNORE_OPTIONS[ignore_par[1]] + ' for ' + ignore_par[0])
                return True
        elif ignore_par[1] == 2:
            if link[len(link) - len(ignore_par[0]):] == ignore_par[0]:
                print(link + ' ignored, with option ' + IGNORE_OPTIONS[ignore_par[1]] + ' for ' + ignore_par[0])
                return True
        else:
            print('Unknown ignore option key ' + str(ignore_par[1]))

    return False


def ignore_links(links):
    """
    """
    filtered = []

    for link in links:
        filter_flag = False
        for ignore_par in IGNORE_LIST:
            if ignore_par[1] == 0:
                if link[:len(ignore_par[0])] == ignore_par[0]:
                    print(link + ' ignored, with option ' + IGNORE_OPTIONS[ignore_par[1]] + ' for ' + ignore_par[0])
                    filter_flag = True
                    break
            elif ignore_par[1] == 1:
                if ignore_par[0] in link:
                    print(link + ' ignored, with option ' + IGNORE_OPTIONS[ignore_par[1]] + ' for ' + ignore_par[0])
                    filter_flag = True
                    break
            elif ignore_par[1] == 2:
                if link[len(link) - len(ignore_par[0]):] == ignore_par[0]:
                    print(link + ' ignored, with option ' + IGNORE_OPTIONS[ignore_par[1]] + ' for ' + ignore_par[0])
                    filter_flag = True
                    break
            else:
                print('Unknown ignore option key ' + str(ignore_par[1]))

        if not filter_flag:
            filtered.append(link)

    return filtered


def link_exists(links, link_str):
    """
    """
    for page_links in links:
        for link in page_links:
            if link_str == link:  # to be improved by checking website name or base link instead
                return True
    return False


def remove_duplicate_links(links):
    """
    """
    pass


def get_all_links_on_page(driver, page_link, visited):
    """
    """
    print("Getting all links on page: " + page_link)
    links_on_page = []
    url_obj = urlparse(page_link)
    website = ''
    link_parts = page_link.split('/')
    i = 0
    while (link_parts[i] == 'http:' or link_parts[i] == 'https:' or link_parts[i] == ''):
        i += 1
    if i < len(link_parts):
        website = link_parts[i]
    sleep(randint(1, 200) / 1000.)
    driver.get(page_link)
    soup = BeautifulSoup(driver.page_source, 'lxml')
    link_tags = soup.find_all('a')
    for link_tag in link_tags:
        try:
            link_str = link_tag['href']
            if link_str not in page_links and not ignore_single_link(link_str):
                if link_str[0] == '/':
                    new_link = url_obj.netloc + link_str
                    if new_link not in visited:
                        links_on_page.append(new_link)
                        visited.add(new_link)
                    new_link = website + link_str
                    if new_link not in visited:
                        links_on_page.append(new_link)
                        visited.add(new_link)
                elif link_str[0:2] == './':
                    new_link = page_link + link_str[1:]
                    if new_link not in visited:
                        links_on_page.append(new_link)
                        visited.add(new_link)
                    new_link = website + link_str[1:]
                    if new_link not in visited:
                        links_on_page.append(new_link)
                        visited.add(new_link)
                elif link_str[0:4] != 'http':
                    new_link = page_link + '/' + link_str
                    if new_link not in visited:
                        links_on_page.append(new_link)
                        visited.add(new_link)
                    new_link = website + '/' + link_str
                    if new_link not in visited:
                        links_on_page.append(new_link)
                        visited.add(new_link)
                else:
                    if link_str not in visited:
                        links_on_page.append(link_str)
                        visited.add(new_link)
        except:
            continue
    return ignore_links(links_on_page)


def get_links(driver, SEARCH_STRING):
    """
    """
    links = []
    start_depth = DEPTH * 10
    BASE_SEARCH_LINK = GOOG_SEARCH_LINK  # default search engine is Google
    START_STRING = GOOG_START_STRING
    STEP = 10
    visited = set()
    for i in range(0, start_depth, STEP):
        print('Loading results from search page ' + str(i // STEP + 1))
        page_links = []
        SEARCH_LINK = BASE_SEARCH_LINK + adjust_search_string(SEARCH_STRING) + START_STRING + str(i)
        sleep(randint(1, 200) / 1000.)
        driver.get(SEARCH_LINK)
        soup = BeautifulSoup(driver.page_source, 'lxml')
        # print(soup.prettify())
        link_tags = soup.find_all('a')

        for link_tag in link_tags:
            try:
                link_str = link_tag['href']
                if link_str not in visited:
                    if link_str not in page_links and not ignore_single_link(link_str):
                        page_links.append(link_str)
                        visited.add(link_str)
                        page_links += get_all_links_on_page(driver, link_str, visited)
            except:
                continue

        page_links = ignore_links(page_links)

        links.append(page_links)

        print('Number of links: ' + str(len(page_links)))
        # print(links)
    print("Total number of visited links: " + str(len(visited)))
    return links


def find_email_on_page(driver, page_link, visited_emails):
    """
    """
    print("Loading emails on page: " + page_link)
    sleep(randint(1, 200) / 1000.)
    driver.get(page_link)
    soup = BeautifulSoup(driver.page_source, 'lxml')
    regex = r'[^@]+@[^@]+\.[^@]+'

    text = [s.extract() for s in soup(['p', 'span', 'a'])]

    # print(text)
    emails = []
    for split_text in text:

        for word in split_text:
            try:
                ss = BeautifulSoup(word, 'lxml')
                content = str(ss.getText())
                if content is not None and content.strip() != '':
                    # print('Content: '+str(content)) #debug line
                    for w in content.split():
                        if re.match(regex, w) and w not in visited_emails:
                            emails.append(w)
                            visited_emails.add(w)
            except:
                continue
    title = soup.findAll('title')[0].getText().strip()
    print('title: ' + title)
    return emails, title


def flatten(lis):
    """
    """
    print("Flattening emails")
    ss = lis[0]
    for i in range(1, len(lis)):
        ss += ' , ' + lis[i]

    return ss


def export_results(results, keyword):
    """
    """
    wb = Workbook()
    sheet = wb.active
    sheet['A1'] = 'Keyword'
    sheet['B1'] = 'Website'
    sheet['C1'] = 'Page Title'
    sheet['D1'] = 'Full Link'
    sheet['E1'] = 'Search Page Number'
    sheet['F1'] = 'Emails'

    for i in range(len(results)):
        print('Adding line ' + str(i + 2) + ' to export file')
        sheet['A' + str(i + 2)] = results[i]['keyword']
        sheet['B' + str(i + 2)] = results[i]['website']
        sheet['C' + str(i + 2)] = results[i]['title']
        sheet['D' + str(i + 2)] = results[i]['full_link']
        sheet['E' + str(i + 2)] = results[i]['page_number']
        sheet['F' + str(i + 2)] = results[i]['emails']

    """
    counter = 1
    while os.path.exists('export '+str(counter)+'.xlsx'):
        counter+=1
    
    wb.save('export '+str(counter)+'.xlsx')
    """
    tt = datetime.now()
    time_str = str(tt.year) + str(tt.month) + str(tt.day) + str(tt.hour) + str(tt.minute)

    export_filename = 'export_' + keyword + '_' + time_str + '.xlsx'

    try:
        wb.save(export_filename)
        print('exportation to file: ' + export_filename)
    except:
        export_filename = str(randint(1, 1000)) + '_' + export_filename
        wb.save(export_filename)
        print('exportation to file: ' + export_filename)


if __name__ == '__main__':
    """
    """

    driver = set_up_driver()
    # exit()

    program_start = time()

    # load keyword list
    Tk().withdraw()
    keyword_file = askopenfilename()
    keywords = load_keyword_dict(keyword_file)

    # print(keywords) #debug line
    # SEARCH_STRING = input("Enter search string: ")

    print('Keywords loaded')

    # first_results = {}
    final_results = []
    for SEARCH_STRING in keywords.keys():
        if keywords[SEARCH_STRING] != 'DONE':
            keyword_start = time()
            print('Searching for results for keyword: ' + SEARCH_STRING)
            links = get_links(driver, SEARCH_STRING)
            """
            page_id = 1
            for page_links in links:
                first_results[SEARCH_STRING] = {}
                first_results[SEARCH_STRING][page_id] = {}
                first_results[SEARCH_STRING][page_id]['page_links']  = page_links
                page_id+=1
            """
            page_id = 1
            visited_emails = set()
            for page_links in links:
                for link in page_links:
                    try:
                        emails, title = find_email_on_page(driver, link, visited_emails)
                        if len(emails) > 0:
                            final_results.append(
                                {'keyword': SEARCH_STRING, 'website': urlparse(link).netloc, 'title': title,
                                 'full_link': link, 'page_number': page_id, 'emails': flatten(emails)})
                        if len(final_results) > 0 and len(final_results) % 10 == 0:
                            export_results(final_results, SEARCH_STRING)
                            print('temporary file exported')
                    except:
                        continue
                page_id += 1
            sleep(randint(1, 200) / 1000.)

            # print('final_results: '+str(final_results))
            export_results(final_results, SEARCH_STRING)
            print('file exported')
            status_change = set_keyword_status(keyword_file, SEARCH_STRING, 'DONE')
            if not status_change:
                print('WARNING: status could not be changed for keyword ' + SEARCH_STRING)
            keyword_end = time()
            print('Keyword search ended for ' + SEARCH_STRING + '. Total duration: ' + str(
                round(keyword_end - keyword_start, 2)) + ' seconds')

    # emails = find_email_on_page(driver,'https://www.dotdash.com/careers/?p=jobs&nl=1')
    # print(emails)
    driver.quit()
    program_end = time()
    print(BOTNAME+' ended. Total duration: ' + str(round(program_end - program_start, 2)) + ' seconds')
